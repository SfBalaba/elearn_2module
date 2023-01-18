import csv, re, os, datetime
from typing import List, Dict, Tuple, Any

import cProfile
from pstats import Stats, SortKey

import matplotlib
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html

from jinja2 import Environment, FileSystemLoader
import pdfkit

currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }


class DataSet:
    """Класс представляющий данные из файла .csv

        Attributes:
            file_name (str): название вводимого файла csv
            vacancies_objects (List[Vacancy]): список состоящий из экземпляров класса Vacancy,
             содержащий только те строки, которые заполнены полностью, и не содержат пустые элементы.

    """
    def __init__(self, file_name):
        """Инициализирует экземпляр класса DataSet

        Args:
            file_name (str): название вводимого файла csv
            vacancies_objects (List[Vacancy]): список состоящий из экземпляров класса Vacancy,
             содержащий только те строки, которые заполнены полностью, и не содержат пустые элементы.

        >>> type(DataSet('vacancies_by_year.csv')).__name__
        'DataSet'
        >>> DataSet('vacancies_by_year.csv').file_name
        'vacancies_by_year.csv'
        """
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def clean_string(self, raw_html):
        """Очищает строки от html тэгов и лишних пробелов

        Args:
            raw_html (str): строка csv файла
        Returns:
            str: строка без лишних пробелов и html тэгов
        >>> data = DataSet('test.csv')
        >>> data.clean_string("<strong>Программист 1С (8.0)/Разработчик</strong>,65000.0,75000.0,RUR,Москва,2007-12-04T16:28:52+0300")
        'Программист 1С (8.0)/Разработчик,65000.0,75000.0,RUR,Москва,2007-12-04T16:28:52+0300'
        >>> data = DataSet('test.csv')
        >>> data.clean_string("  Программист      1С (8.0)/Разработчик")
        'Программист 1С (8.0)/Разработчик'
        """
        result = re.sub("<.*?>", '', raw_html)
        return result if '\n' in raw_html else " ".join(result.split())

    def csv_reader(self, file_name: str) -> (list, list):
        """
        Считывает записи из файла file_name

        Args:
            file_name (str): название файла csv

        Returns:
            Tuple(List[str], List[str]): (список заголовков CSV и
            список списков, где каждый вложенный список содержит все поля вакансии)
        >>> data.csv_reader(data.file_name)[0]
         ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
        >>> data.csv_reader(data.file_name)[1]
        [[],
         ['Программист баз данных', '36000.0', '50000.0', 'FALSE', 'RUR', 'Москва', '2007-12-04T11:27:27+0300'],
          ['<strong>Программист 1С (8.0)/Разработчик</strong>', '65000.0', '75000.0', 'RUR', 'Москва', '2007-12-04T16:28:52+0300']]

        """
        reader = csv.reader(open(file_name, encoding='utf_8_sig'))
        data_base = [line for line in reader]
        return data_base[0], data_base[1:]

    def csv_filer(self, list_naming: list, reader: list) -> list:
        """Отсекает те строки, которые заполнены не полностью, или содержат пустые элементы
         и возвращает список словарей, представляющих вакансию, где ключи содержат все поля вакансии
                
        Args:
            list_naming (list): список названий полей 
            reader (List[str]): список списков, где каждый вложенный список содержит все поля вакансии
            
        Returns:
            List[Dict[str, str]]: список словарей, представляющих вакансию, где ключи содержат все поля вакансии

        >>> header = ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
        >>> reader = [[]]
        >>> data.csv_filer(header, reader)
        []
        >>> header = ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
        >>> reader1 = [['<strong>Программист 1С (8.0)/Разработчик</strong>', '65000.0', '75000.0', 'RUR', 'Москва', '2007-12-04T16:28:52+0300']]
        >>> data.csv_filer(header, reader1)
        [{'name': 'Программист 1С (8.0)/Разработчик', 'salary_from': '65000.0', 'salary_to': '75000.0', 'salary_currency': 'RUR', 'area_name': 'Москва', 'published_at': '2007-12-04T16:28:52+0300'}]
        >>> reader =  [['Веб-программист','30000.0','','RUR','Москва','2007-12-04T15:06:07+0300']]
        >>> data.csv_filer(header, reader)
        []
        >>> reader =  [['Веб-программист','30000.0','RUR','Москва','2007-12-04T15:06:07+0300']]
        >>> data.csv_filer(header, reader)
        []
        """
        new_vacans_list = list(filter(lambda vac: (len(vac) == len(list_naming) and vac.count('') == 0), reader))
        dict_vacans = [dict(zip(list_naming, map(self.clean_string, vac))) for vac in new_vacans_list]
        return dict_vacans



class Vacancy:
    """Класс представляющий экземпляр вакансии

    Attributes:
        name (str): название вакансии
        salary (Salary): объект зарплата, содержит информацию о вилке оклада и валюте
        area_name (str): город вакансии
        published_at (int): год публикации
    """
    def __init__(self, dict_vac) -> object:
        """Инициализирует объект Vacancy

        Args:
            dict_vac (dict(str, str)): словарь, представляющий вакансию,
             где ключи содержат все поля вакансии
        >>> test = {'name':'Программист баз данных','salary_from':'36000.0','salary_to': '50000.0','salary_currency': 'RUR','area_name':'Москва','published_at':'2007-12-04T11:27:27+0300'}
        >>> type(Vacancy(test)).__name__
        'Vacancy'
        >>> Vacancy(test).name
        'Программист баз данных'
        >>> Vacancy(test).area_name
        'Москва'
        >>> Vacancy(test).salary.__class__.__name__
        'Salary'
        >>> (Vacancy(test).published_at).__class__.__name__
        'int'
        >>> Vacancy(test).published_at
        2007
        """
        self.name = dict_vac['name']
        self.salary = Salary(dict_vac['salary_from'], dict_vac['salary_to'], dict_vac['salary_currency'])
        self.area_name = dict_vac['area_name']
        self.published_at = self.change_data(dict_vac['published_at'])

    def change_data(self, date_vac):
        """Приводит дату публикаци вакансии в формат гггг

        Returns:
            int: год публикации вакансии

        >>> test = {'name':'Программист баз данных','salary_from':'36000.0','salary_to': '50000.0','salary_currency': 'RUR','area_name':'Москва','published_at':'2008-12-04T11:27:27+0300'}
        >>> vac = Vacancy(test)
        >>> vac.change_data(test['published_at'])
        2008

        >>> result = vac.change_data('2008-12-04T11:27:27+0300')

        >>> result.__class__.__name__
        'int'
        """
        if date_vac.__class__.__name__ != 'str':
            raise TypeError('Argument must be string type')
        else:
            return int(datetime.datetime.strptime(date_vac, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            # return int(date_vac.split('-').[0])
            # return int(date_vac[:4])


class Salary:
    """Класс для представления зарплаты

        Attributes:
            salary_from (str): верхняя граница вилки оклада
            salary_to (str):нижняя граница вилки оклада
            salary_currency (str): валюта оклада

        """
    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализирует объект Salary

        Args:
            salary_from (str or int or float): верхняя граница вилки оклада
            salary_to (str or int or float):нижняя граница вилки оклада
            salary_currency (str): валюта оклада
    >>> type(Salary(20000, 40000, 'RUR')).__name__
    'Salary'
    >>> Salary(20000, 40000, 'RUR').salary_to
    40000
    >>> Salary(20000, 40000, 'RUR').salary_from
    20000
    >>> Salary(20000, 40000, 'RUR').salary_currency
    'RUR'
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def convert_to_RUB(self) -> float:
        """Конвертирует оклад в рубли

        Args:
            salary: среднее арифметическое между salary_from и salary_to(оклад)

        Returns:
            оклад в рублях
        >>> Salary(1000, 5000, 'EUR').convert_to_RUB()
        179700.0
        >>> Salary(1000, 5000, 'KGS').convert_to_RUB()
        2280.0
        >>> Salary(1000, 5000, 'RUR').convert_to_RUB()
        3000.0
        >>> Salary(1000.5, 5000, 'EUR').convert_to_RUB()
        179714.975

        """

        return (float(self.salary_from) + float(self.salary_to)) / 2 * currency_to_rub[self.salary_currency]


def get_dynamic_by_salary(vacancies: List[Vacancy], field: str, filter_name_vacancy: str = ''):
    """Представляет динамику уровня зарплат по заданному полю для названия вакансии(необязательное поле)

    Attributes:
        vacancies (List[Vacancy]): список экземпляров объектов класса Vacancy
        field (str): название поля, по которому производится динамика
        filter_name_vacancy (str): название профессий для которой производится обработка статистических данных

    Returns:
        Dict(str or int: int): словарь, подставляющий стратистику ключ-значение поля(field),
         значение: среднее арифметическое зарплат вакансий по заданному значению названия поля

    """
    by_salary = {}
    for vac in vacancies:
        by_salary[vac.__getattribute__(field)] = [] if vac.__getattribute__(field) not in by_salary.keys() else \
            by_salary[
                vac.__getattribute__(field)]
    if len(filter_name_vacancy) != 0:
        for vac_name in filter_name_vacancy:
            vacancies += list(filter(lambda vac: vac_name in vac.__getattribute__('name'), vacancies))

    for vac in vacancies:
        by_salary[vac.__getattribute__(field)].append(
            vac.salary.convert_to_RUB())
    for key in by_salary.keys():
        by_salary[key] = 0 if len(by_salary[key]) == 0 else int(sum(by_salary[key]) // len(by_salary[key]))
    return by_salary


def get_dynamic_by_count(vacancies: List[Vacancy], field: str, filter_name_vacancy: str = ''):
    """Представляет динамику количества вакансий по заданному полю для вакансии по названию(необязательное поле)

    Attributes:
        vacancies (List[Vacancy]): список экземпляров объектов класса Vacancy
        field (str): название поля, по которому производится динамика
        filter_name_vacancy (str): название профессий для которой производится обработка статистических данных

    Returns:
        dict(str or int: int): словарь, подставляющий стратистику ключи-все возможные значение поля(field),
         значение: количество вакансий по заданному значению названия поля,
          для 'area_name'(город)-доля вакансий
    """
    by_count = {}
    for vac in vacancies:
        by_count[vac.__getattribute__(field)] = 0 if vac.__getattribute__(field) not in by_count.keys() else by_count[
            vac.__getattribute__(field)]
    new_list= []
    if len(filter_name_vacancy) != 0:
        for vac_name in filter_name_vacancy:
            new_list = list(filter(lambda vac: vac_name in vac.__getattribute__('name'), vacancies))
            vacancies += new_list

    for vac in vacancies:
        by_count[vac.__getattribute__(field)] += 1
    if field == 'area_name':
        for key in by_count.keys():
            by_count[key] = round(by_count[key] / len(data.vacancies_objects), 4)
    return by_count


def exit_from_file(message):
    """Принудительно завершает выполнение программы и выводит сообщение о некорректном пользовательском вводе

    Args:
        message (str): сообщение о некорректном пользовательском вводе
    """
    print(message)
    exit()


class Report:
    """Класс для представления отчётов по статистике

    Attributes:
            vacancy (str): название вакансии для которой предоставляется отчёт
            salary_by_year (dict(int: int)): Динамика уровня зарплат по годам
            salary_by_year_vac (dict(int: int)): Динамика уровня зарплат по годам для выбранной профессии(vacancy)
            count_by_year (dict(int: int)): Динамика количества вакансий по годам
            count_by_year_vac (dict(int: int)): дДинамика количества вакансий по годам для выбранной профессии(vacancy)
            salary_by_city (dict(str: int)): Уровень зарплат по городам (в порядке убывания) - только первые 10 значений
            pers_by_city (dict(str: int)): Доля вакансий по городам (в порядке убывания) - только первые 10 значений
            statistics (List(dict(int or str: int))): список собранных статистических данных,
             исползуемых для составления табилц и графиков
    """
    def __init__(self, salary_by_year: Dict, salary_by_year_vac: Dict, count_by_year: Dict, count_by_year_vac: Dict,
                 salary_by_city: Dict,
                 pers_by_city: Dict, vac=""):
        """Инициализирует объект report для создания отчёта в pdf, содержащий таблицы и графики
         с статистикой по годам и городам

        Args:
            vac (str): название вакансии для которой предоставляется отчёт(необязательное поле)
            salary_by_year (dict(int: int)): Динамика уровня зарплат по годам
            salary_by_year_vac (dict(int: int)): Динамика уровня зарплат по годам для выбранной профессии(vacancy)
            count_by_year (dict(int: int)): Динамика количества вакансий по годам
            count_by_year_vac (dict(int: int)): дДинамика количества вакансий по годам для выбранной профессии(vacancy)
            salary_by_city (dict(str: int)): Уровень зарплат по городам (в порядке убывания) - только первые 10 значений
            pers_by_city (dict(str: int)): Доля вакансий по городам (в порядке убывания) - только первые 10 значений
        """
        self.vacancy = vac
        self.salary_by_year = salary_by_year
        self.salary_by_year_vac = salary_by_year_vac
        self.count_by_year = count_by_year
        self.count_by_year_vac = count_by_year_vac
        self.salary_by_city = salary_by_city
        self.pers_by_city = pers_by_city
        self.statistics = [salary_by_year, salary_by_year_vac, count_by_year, count_by_year_vac, salary_by_city,
                           self.pers_by_city]

    def аgenerate_image(self):
        """Генерирует изображение в директории report под именем report.png, с 4мя графиками: гистограммами 'Уровень зарплат по годам',
         'Количество вакансий по годам', 'Уровень зарплат по городам' и круговой диаграммой 'Количество вакансий по городам'
          для выбранной профессии, сохраняет сгенерированное изображение в корень проекта с названием grapth.png

        """
        labels = self.salary_by_year.keys()
        vacancy = self.vacancy

        x = np.arange(len(labels))
        width = 0.4
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        ax1.bar(x - width / 2, self.salary_by_year.values(), width, label='средняя з/п')
        ax1.bar(x + width / 2, self.salary_by_year_vac.values(), width, label=f'з/п {vacancy[2]}')
        ax1.set_title('Уровень зарплат по годам')
        plt.sca(ax1)
        plt.xticks(x, labels, fontsize=8, rotation=90)

        ax1.legend(loc='upper left', fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, self.count_by_year.values(), width, label='Количество вакансий')
        ax2.bar(x + width / 2, self.count_by_year_vac.values(), width, label=f'Количество вакансий \n{vacancy[2]}')
        ax2.set_title('Количество вакансий по годам')
        plt.sca(ax2)
        plt.xticks(x, labels, fontsize=8, rotation=90)

        ax2.legend(loc='upper left', fontsize=8)
        ax2.grid(axis='y')

        cities = self.salary_by_city.keys()
        y_pos = np.arange(len(cities))
        ax3.barh(y_pos, self.salary_by_city.values(), align='center')
        plt.sca(ax3)
        plt.yticks(y_pos, cities, fontsize=6)
        ax3.invert_yaxis()  # labels read top-to-bottom
        ax3.set_title('Уровень зарплат по городам')
        ax3.grid(axis='x')

        x = ['Другие'] + list(self.pers_by_city.keys())
        ax4.set_title('Доля вакансий по городам')
        city_percent = list(self.pers_by_city.values())
        city_percent.insert(0, 1 - sum(city_percent))
        ax4.pie(city_percent, radius=1, labels=x, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('report/graph.png')

    def generate_exel(self):
        """Генерирует xlsx файл в директории report под именем report.xlsx с двумя вкладками:
         "Статистика по годам", "Статистика по городам",
         содержащая общую статистику и для укаазанной в отдельности
            Args:
                list_vacancies (list(dict(str or int, int))): список различных статистических данных по годам и городам
                 с динамикой долей вакансий и уровнем зарплат
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws2 = wb.create_sheet("Статистика по городам")
        data_1 = self.statistics[:4]
        data_2 = self.statistics[4:5]
        data_3 = self.statistics[5:]
        ws1.append(['Год', "Средняя зарплата", f"Средняя зарплата - {self.vacancy[2]}",
                    "Количество вакансий", f"Количество вакансий - {self.vacancy[2]}"])
        ws2.append(["Город", "Уровень зарплат", "", "Город", "Доля зарплат"])
        keys = list(data_1[0].keys())
        self.create_table(ws1, keys, data_1, min_c=1, min_r=2, max_c=5)
        self.create_table(ws2, list(data_2[0].keys()), data_2, min_c=1,
                          min_r=2,
                          max_c=2)
        self.create_table(ws2, list(data_3[0].keys()), data_3, min_c=4, min_r=2,
                          max_c=5)
        self.set_style(wb)

        file_name = os.path.join(os.getcwd(), 'report/report.xlsx')
        if (isinstance(file_name, str)):
            if (os.path.basename(file_name).endswith(".xlsx")):
                if (os.path.exists(os.path.dirname(file_name))):
                    wb.save("report/report.xlsx")
                else:
                    raise FileNotFoundError('Такой директории не существует')
            else:
                raise "Невреное расширение файла"
        else:
            raise TypeError('Имя файла должно иметь тип str')

    def create_table(self, ws: openpyxl.worksheet, keys: List, data_table: List[Dict[str,str]], min_c: int, min_r: int,
                     max_c: int):
        """Создаёт таблицу во вкладке ws с аттрибутами keys, значениями data_table,
         в диапазоне стобцов min_c, max_c, начиная со строки min_r
        Args:
             ws (openpyxl.worksheet): текущая вкладка в которую идёт запись
             keys (List(str)): название аттрибутов новой таблицы
             data_table (List[Dict[str,str]]): необходимые статистические данные, которыми заполняется таблица
             позиция ячейки  с которой начинается заполнение:
             min_c (int): номер колонки начальной позиции
             min_r (int): номер строки начальной позиции
             max_c (int): столбец которым заканчивается заполнение
        """
        for col in ws.iter_cols(min_col=min_c, min_row=min_r, max_col=max_c, max_row=len(keys) + 1):
            for cell in col:
                row_number = cell.row
                if (cell.column == min_c):
                    cell.value = keys[row_number - min_r]
                else:
                    try:
                        current_key = int(ws.cell(row=row_number, column=min_c).value)
                    except ValueError:
                        current_key = ws.cell(row=row_number, column=min_c).value
                    curr_statistic = data_table[cell.column - min_c - 1]
                    cell.value = curr_statistic[current_key]


    def set_style(self, wb: openpyxl.worksheet):
        """Задаёт стиль для рабочей книги wb: обводка границ ячеек и жирное начертание для шапки,
         статистика количества вакансий по городам в процентном соотношении до 2-ух знаков после точки

        Args:
             wb (openpyxl.worksheet): рабочая книга для которой задаётся стиль
        """
        thins = Side(border_style="thin")
        names = wb.sheetnames
        max_widths = {}
        for name in names:
            curr_ws = wb[name]
            col_width = []
            for col in curr_ws.columns:
                widths = []
                for cell in col:

                    if cell.value != None and cell.value != "":
                        cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        if (cell.row == 1):
                            cell.font = Font(bold=True)
                    widths.append(len(str(cell.value))) if (cell.value != None) else widths.append(0)
                col_width.append(max(widths))
            max_widths[name] = col_width

        for i in range(2, 12):
            wb["Статистика по городам"][f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for name in names:
            for i in range(wb[name].max_column):
                wb[name].column_dimensions[get_column_letter(i + 1)].width = max_widths[name][i] + 1.22


    def generate_pdf(self) -> None:
        """Генерирует pdf в директории report под именем report.pdf, который содержит грфики(generate_image) и таблицы(generate_exel)
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report/report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()
        out2 = xlsx2html('report/report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'name_vacancy': self.vacancy, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report/report.pdf', configuration=config,
                           options={"enable-local-file-access": ""})


def print_statistic(result_list, index, message, is_reversed=False, slice=0):
    """Выводит собранную форматированные статистику
      Вывод в виде: "<сообщение о параметрах статистики>: <статистические данные>(result_list)"

    Args:
        result_list (dict_items(str or int, int)): статистические данные
        index (int): параметр, по которому происходит сортировка 0-по ключчу, 1-по значению
        message (str): сообщение, содержащее информацию о статистике, которая выводится в текущий момент
        is_reversed (bool): порядок сортировки False-в порядке возрастания, True-в порядке убывания
        slice (int): количесвто выводимых записей статистики, если параметр не указан, то выводится весь список
    """
    slice = len(result_list) if slice == 0 else slice
    statistic = dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])
    print(message + str(statistic))


def get_statistic(result_list, index: int, is_reversed=False, slice=0):
    """Возвращает форматированные статистические данные

    Args:
        result_list (dict_items(str or int, int)): статистические данные
        index (int): параметр, по которому происходит сортировка 0-по ключчу, 1-по значению
        is_reversed (bool): порядок сортировки False-в порядке возрастания, True-в порядке убывания
        slice (int): количесвто выводимых записей статистики, если параметр не указан, то выводится весь список

    Returns:
         dict(int or str, int): статистические данные заданной длинны(slice), если параметр не указан,
    то выводится весь список, предварительно сортируя в порядке(is_reversed)
    """
    slice = len(result_list) if slice == 0 else slice
    statistic = dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])
    return statistic

# file_name = input('Введите название файла: ')
# vacancy_name = input('Введите название профессии: ')
# if os.stat(file_name).st_size == 0:
#     exit_from_file('Пустой файл')
# data = DataSet(file_name)
# if len(data.vacancies_objects) == 0:
#     exit_from_file('Нет данных')
# dict_cities = {}

def get_all_stat(file_name, vacancy_name):
    global vac, pers_by_city, data, pers_by_city
    data = DataSet(file_name)
    if len(data.vacancies_objects) == 0:
        exit_from_file('Нет данных')
    dict_cities = {}
    for vac in data.vacancies_objects:
        if vac.area_name not in dict_cities.keys():
            dict_cities[vac.area_name] = 0
        dict_cities[vac.area_name] += 1
    needed_vacancies_objects = list(
        filter(lambda vac: int(len(data.vacancies_objects) * 0.01) <= dict_cities[vac.area_name],
               data.vacancies_objects))
    print_statistic(get_dynamic_by_salary(data.vacancies_objects, 'published_at').items(), 0,
                    'Динамика уровня зарплат по годам: ')
    print_statistic(get_dynamic_by_count(data.vacancies_objects, "published_at").items(), 0,
                    'Динамика количества вакансий по годам: ')
    print_statistic(get_dynamic_by_salary(data.vacancies_objects, 'published_at', vacancy_name).items(), 0,
                    'Динамика уровня зарплат по годам для выбранной профессии: ')
    print_statistic(get_dynamic_by_count(data.vacancies_objects, 'published_at', vacancy_name).items(), 0,
                    'Динамика количества вакансий по годам для выбранной профессии: ')
    print_statistic(get_dynamic_by_salary(needed_vacancies_objects, 'area_name').items(), 1,
                    'Уровень зарплат по городам (в порядке убывания): ', True, 10)
    print_statistic(get_dynamic_by_count(needed_vacancies_objects, 'area_name').items(), 1,
                    'Доля вакансий по городам (в порядке убывания): ', True, 10)
    salary_by_year = get_statistic(get_dynamic_by_salary(data.vacancies_objects, 'published_at').items(), 0)
    count_by_year = get_statistic(get_dynamic_by_count(data.vacancies_objects, "published_at").items(), 0)
    salary_by_year_vac = get_statistic(
        get_dynamic_by_salary(data.vacancies_objects, 'published_at', vacancy_name).items(), 0)
    count_by_year_vac = get_statistic(
        get_dynamic_by_count(data.vacancies_objects, 'published_at', vacancy_name).items(),
        0)
    salary_by_city = get_statistic(get_dynamic_by_salary(needed_vacancies_objects, 'area_name').items(), 1, False,
                                   slice=10)
    pers_by_city = get_statistic(get_dynamic_by_count(needed_vacancies_objects, 'area_name').items(), 1, True, slice=10)

    return salary_by_year, count_by_year,salary_by_year_vac,count_by_year_vac,salary_by_city,pers_by_city


# salary_by_year, count_by_year,salary_by_year_vac,count_by_year_vac,salary_by_city,pers_by_city = get_all_stat(file_name)
# report = Report(salary_by_year=salary_by_year,
#                 salary_by_year_vac=salary_by_year_vac,
#                 count_by_year=count_by_year,
#                 count_by_year_vac=count_by_year_vac,
#                 salary_by_city=salary_by_city,
#                 pers_by_city=pers_by_city,
#                 vac=vacancy_name)
# report.generate_exel()
# report.generate_image()
# report.generate_pdf()

if __name__ == "__main__":
    vacancy_name = ['web develop', 'веб разработчик', 'web разработчик', 'web programmer', 'web программист', 'веб программист', 'битрикс разработчик', 'bitrix разработчик', 'drupal разработчик', 'cms разработчик', 'wordpress разработчик', 'wp разработчик', 'joomla разработчик', 'drupal developer', 'cms developer', 'wordpress developer', 'wp developer', 'joomla developer']
    file_name = 'vacancies_with_skills.csv'
    salary_by_year, count_by_year,salary_by_year_vac,count_by_year_vac,salary_by_city,pers_by_city = get_all_stat(file_name, vacancy_name)
    report = Report(salary_by_year=salary_by_year,
                    salary_by_year_vac=salary_by_year_vac,
                    count_by_year=count_by_year,
                    count_by_year_vac=count_by_year_vac,
                    salary_by_city=salary_by_city,
                    pers_by_city=pers_by_city,
                    vac=vacancy_name)
    report.generate_exel()
    report.generate_image()
    report.generate_pdf()
